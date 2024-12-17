import multiprocessing
import os
import threading
import time
import PyPDF2
from concurrent.futures import ProcessPoolExecutor
import queue 
import pdfplumber
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

# Función para dividir el PDF en subarchivos más pequeños
def dividir_pdf(pdf_path, num_parts):
    start_time = time.time()
    print("Inicio de división de PDF")
    
    pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
    output_paths = []
    with open(pdf_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        total_pages = len(reader.pages)

        pages_per_part = total_pages // num_parts
        remainder = total_pages % num_parts

        start_page = 0
        for i in range(num_parts):
            end_page = start_page + pages_per_part + (1 if i < remainder else 0)
            output_path = f"{pdf_name}_part_{i + 1}.pdf"
            writer = PyPDF2.PdfWriter()

            for page_num in range(start_page, end_page):
                writer.add_page(reader.pages[page_num])

            with open(output_path, "wb") as output_pdf:
                writer.write(output_pdf)

            output_paths.append(output_path)
            start_page = end_page
    
    end_time = time.time()
    print(f"División de PDF completada en {end_time - start_time:.2f} segundos.")
    return output_paths, total_pages  # También devolvemos el número total de páginas

def monitor_progress_queue(progress_queue, total_pages, progress_callback, stop_event):
    pages_processed = 0
    while not stop_event.is_set() or not progress_queue.empty():
        try:
            increment = progress_queue.get(timeout=0.1)
            pages_processed += increment
            if progress_callback:
                progress_callback(pages_processed, total_pages)
        except queue.Empty:
            continue
    # Actualización final
    if progress_callback:
        progress_callback(pages_processed, total_pages)

# Función para extraer la cabecera y la tabla
def extraer_cabecera_y_tabla(bloque):
    lineas = bloque.split("\n")
    cabecera = {}
    for i, linea in enumerate(lineas):
        if "Documento" in linea:
            doc = " ".join((linea.split(":")[-4].strip()).split()[:-1])
            cabecera['Documento'] = doc
        if "Rol" in linea:
            rol = " ".join((linea.split(":")[-3].strip()).split()[:-1])
            cabecera['Rol'] = rol
        if "Nombre" in linea:
            nombre = " ".join((linea.split(":")[-2].strip()).split()[:-1])
            cabecera['Nombre'] = nombre
        if "CDepto" in linea:
            cdepto = linea.split(":")[-1].strip()
            cod_cdepto = cdepto.split(" ")[0]
            cdepto = " ".join(cdepto.split(" ")[1:])
            cabecera['CDepto'] = cdepto
        if "Cargo" in linea:
            cargo = " ".join((linea.split(":")[-2].strip()).split()[:-1])
            cod_cargo = cargo.split(" ")[0]
            cargo = " ".join(cargo.split(" ")[1:])
            cabecera['Cargo'] = cargo
        if "CCosto" in linea:
            ccosto = linea.split(":")[-1].strip()
            cod_ccosto = ccosto.split(" ")[0]
            ccosto = " ".join(ccosto.split(" ")[1:])
            cabecera['CCosto'] = ccosto
        if "Fecha" in linea:
            tabla = lineas[i+1:]
            break
    return cabecera, tabla

def encontrar_y_total_empleado(page):
    for char in page.extract_words():
        if "Empleado" in char['text']:
            return (char['top'] - 15)
    return None

def procesar_tabla(page, cabecera, tabla):
    registros = []
    anchos_columnas = [
        38.2, 26.63, 26.9, 29.05, 26.9, 24.48, 26.09, 21.52, 20.18, 21.52,
        20.98, 19.64, 20.28, 16.57, 19.1, 24.48, 20.98, 20.44, 22.87, 20.98,
        23.13, 23.13, 21.79
    ]
    headers = ["Fecha", "Dia", "Tipo", "Programada Entrada", "Programada Salida",
               "Hrs. Prg.", "Hrs. Ref.", "Marca Ing", "Marca Inicio Ref.",
               "Marca Termino Ref.", "Marca Sal", "Hrs. Perm.", "Hrs. Ref. Real",
               "Ind. Err.", "Horas Trabaj.", "Hr. Ext. 1.25", "Hr. Ext. 1.35",
               "Hrs. Dobles", "Hrs. Aus.", "Hrs. No Trab.", "Hr. Ext. Feriado",
               "Hrs. Comp/HP", "Ind. Evento"]

    # Calcular posiciones X de las columnas
    x_positions = [0] + [sum(anchos_columnas[:i+1]) for i in range(len(anchos_columnas))]
    columns = [(x_positions[i], x_positions[i+1]) for i in range(len(x_positions)-1)]

    y0, y1 = 161.5, 600
    y_total_empleado = encontrar_y_total_empleado(page)
    if y_total_empleado:
        y1 = y_total_empleado

    # Extraer palabras dentro del cuadro delimitador
    words = page.extract_words()
    words_in_bbox = [w for w in words if y0 <= w['top'] <= y1]

    # Agrupar palabras en filas basadas en sus posiciones 'top'
    from pdfplumber.utils import cluster_objects
    rows_of_words = cluster_objects(words_in_bbox, 'top', tolerance=1.0)

    for row_words in rows_of_words:
        # Ordenar palabras por posición X
        row_words.sort(key=lambda w: w['x0'])
        row_data = [''] * len(columns)
        for w in row_words:
            x0, x1 = w['x0'], w['x1']
            # Encontrar el índice de la columna
            col_idx = None
            max_overlap = 0
            for i, (col_x0, col_x1) in enumerate(columns):
                overlap = min(x1, col_x1) - max(x0, col_x0)
                if overlap > max_overlap and overlap > 0:
                    max_overlap = overlap
                    col_idx = i
            if col_idx is not None:
                # Agregar o concatenar el texto si hay múltiples palabras en la misma columna
                if row_data[col_idx]:
                    row_data[col_idx] += ' ' + w['text']
                else:
                    row_data[col_idx] = w['text']
        # Crear el registro y agregarlo a la lista
        registro = {**cabecera, **dict(zip(headers, row_data))}
        registros.append(registro)

    return registros

# Función para procesar cada página del PDF (adaptado a tu lógica)
def procesar_pagina(page):
    data_personas = []
    text = page.extract_text()
    bloques = text.split("Libro de Asistencia Individual")
    for bloque in bloques[1:]:
        cabecera, tabla = extraer_cabecera_y_tabla(bloque)
        registros = procesar_tabla(page, cabecera, tabla)
        data_personas.extend(registros)
    return data_personas

# Función para procesar cada parte del PDF
def procesar_parte_pdf(parte_pdf_path, progress_queue):
    start_time = time.time()
    print(f"Inicio de procesamiento de parte PDF: {parte_pdf_path}")
    
    data = []
    with pdfplumber.open(parte_pdf_path) as pdf:
        total_pages_in_part = len(pdf.pages)
        for page_num in range(total_pages_in_part):
            page = pdf.pages[page_num]
            result = procesar_pagina(page)
            data.extend(result)
            progress_queue.put(1)
    
    end_time = time.time()
    print(f"Procesamiento de parte PDF {parte_pdf_path} completado en {end_time - start_time:.2f} segundos.")
    return data

# Función principal para dividir y procesar el PDF en paralelo
def extraer_datos_pdf_parallel(pdf_path, num_workers=10, progress_callback=None):
    print("Inicio del proceso de extracción de datos PDF")
    
    start_time = time.time()

    # Dividir el PDF en partes más pequeñas
    partes_pdf, total_pages = dividir_pdf(pdf_path, num_workers)
    
    data = []
    manager = multiprocessing.Manager()
    progress_queue = manager.Queue()
    
    stop_event = threading.Event()
    progress_thread = threading.Thread(target=monitor_progress_queue, args=(progress_queue, total_pages, progress_callback, stop_event))
    progress_thread.start()
    
    with ProcessPoolExecutor(max_workers=num_workers) as executor:
        futures = [executor.submit(procesar_parte_pdf, parte_pdf, progress_queue) for parte_pdf in partes_pdf]
        for future in futures:
            data.extend(future.result())
    
    stop_event.set()
    progress_thread.join()
    
    for parte_pdf in partes_pdf:
        os.remove(parte_pdf)
    
    end_time = time.time()
    print(f"Proceso de extracción de datos PDF completado en {end_time - start_time:.2f} segundos.")
    
    return data

def agregar_columna_semana(df):
    total_semana_mask = df.apply(lambda row: row.astype(str).str.contains('Total Semana').any(), axis=1)

    df['Total_Semana'] = total_semana_mask
    df['Semana'] = df.groupby('Nombre')['Total_Semana'].cumsum()
    df['Semana'] = df['Semana'] + 1
    df['Semana'] = 'Semana ' + df['Semana'].astype(int).astype(str)
    df.drop('Total_Semana', axis=1, inplace=True)

    cols = df.columns.tolist()
    idx_fecha = cols.index('Fecha')
    cols.insert(idx_fecha + 1, cols.pop(cols.index('Semana')))
    df = df[cols]

    return df

def agregar_columna_tipo_jornada(df_pdf, excel_path):
    df_excel = pd.read_excel(excel_path, dtype = str)
    
    tipo_jornada_map = {
        '048': 'Full Time',
        '019': 'Part Time',
        '030': 'Practicante'
    }

    # Normalizar
    df_pdf['Rol'] = df_pdf['Rol'].astype(str).str.strip()
    df_excel['Codigo Empleado'] = df_excel['Codigo Empleado'].astype(str).str.strip()
    df_excel['Tipo Jornada'] = df_excel['Tipo Jornada'].astype(str).str.strip()

    df_pdf['Tipo Jornada'] = (
        df_pdf['Rol']
        .map(df_excel.set_index('Codigo Empleado')['Tipo Jornada'])  # Mapear usando Codigo Empleado
        .map(tipo_jornada_map)  # Mapear al tipo de jornada
        .fillna('-')  # Rellenar con 'Desconocido' si no hay coincidencias
    )

    # Reordenar las columnas para que Tipo Jornada esté al lado de Rol
    columnas = list(df_pdf.columns)
    indice_rol = columnas.index('Rol')
    columnas.insert(indice_rol + 1, columnas.pop(columnas.index('Tipo Jornada')))
    df_pdf = df_pdf[columnas]

    return df_pdf

def extraer_resumen_semanal(df):
    resumen_mask = df.apply(lambda row: row.astype(str).str.contains('Total Semana').any(), axis=1)
    resumen_semanal = df[resumen_mask].copy()
    df_sin_resumen = df[~resumen_mask].copy()
    
    columnas_a_eliminar = ["Fecha", "Dia", "Tipo", "Programada Entrada", "Programada Salida",
                           "Hrs. Ref.", "Marca Ing", "Marca Inicio Ref.", "Marca Termino Ref.",
                           "Marca Sal", "Ind. Err.", "Ind. Evento"]
    resumen_semanal.drop(columns=columnas_a_eliminar, inplace=True, errors='ignore')
    
    return df_sin_resumen, resumen_semanal

def procesar_saldo_parte_pdf(parte_pdf_path, progress_queue):
    saldo_horas = []
    columnas_saldo_horas = [
        "Hr. Ext. 1.25 - Anterior", "Hr. Ext. 1.35 - Anterior", "Hrs. Dobles - Anterior",
        "Hr. Ext. Feriado - Anterior", "Hr. Ext. 1.25 - Actual", "Hr. Ext. 1.35 - Actual",
        "Hrs. Dobles - Actual", "Hr. Ext. Feriado - Actual"
    ]
    columnas_cabecera = ["Documento", "Rol", "Nombre", "CDepto", "Cargo", "CCosto"]
    
    with pdfplumber.open(parte_pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                bloques = text.split("Libro de Asistencia Individual")
                for bloque in bloques[1:]:
                    cabecera, _ = extraer_cabecera_y_tabla(bloque)
                    lines = bloque.split('\n')
                    for i, line in enumerate(lines):
                        if line.startswith("PE202"):
                            if i + 1 < len(lines):
                                siguiente_fila = lines[i + 1].split()
                                if len(siguiente_fila) == 8:
                                    registro = {**cabecera, **dict(zip(columnas_saldo_horas, siguiente_fila))}
                                    saldo_horas.append(registro)
                            else:
                                print("No hay más filas en esta página.")
    
    progress_queue.put(1)
    return saldo_horas

def extraer_saldo_horas_parallel(pdf_path, num_workers=4, progress_callback=None):
    print("Inicio del proceso de extracción de Saldo de Horas en paralelo")

    start_time = time.time()
    partes_pdf, total_pages = dividir_pdf(pdf_path, num_workers)
    
    saldo_horas = []
    manager = multiprocessing.Manager()
    progress_queue = manager.Queue()
    
    stop_event = threading.Event()
    progress_thread = threading.Thread(target=monitor_progress_queue, args=(progress_queue, total_pages, progress_callback, stop_event))
    progress_thread.start()
    
    with ProcessPoolExecutor(max_workers=num_workers) as executor:
        futures = [executor.submit(procesar_saldo_parte_pdf, parte_pdf, progress_queue) for parte_pdf in partes_pdf]
        for future in futures:
            saldo_horas.extend(future.result())
    
    stop_event.set()
    progress_thread.join()
    
    for parte_pdf in partes_pdf:
        os.remove(parte_pdf)

    end_time = time.time()
    print(f"Proceso de extracción de Saldo de Horas completado en {end_time - start_time:.2f} segundos.")
    
    if saldo_horas:
        columnas_saldo_horas = [
            "Hr. Ext. 1.25 - Anterior", "Hr. Ext. 1.35 - Anterior", "Hrs. Dobles - Anterior",
            "Hr. Ext. Feriado - Anterior", "Hr. Ext. 1.25 - Actual", "Hr. Ext. 1.35 - Actual",
            "Hrs. Dobles - Actual", "Hr. Ext. Feriado - Actual"
        ]
        columnas_cabecera = ["Documento", "Rol", "Nombre", "CDepto", "Cargo", "CCosto"]
        return pd.DataFrame(saldo_horas, columns=columnas_cabecera + columnas_saldo_horas)
    else:
        return None

def tiempo_a_minutos(tiempo_str):
    try:
        if tiempo_str is None or tiempo_str == "":  # Verifica si el tiempo es None o está vacío
            return None

        horas, minutos = map(int, tiempo_str.split(':'))
        return horas * 60 + minutos
    except ValueError:
        return None

def ajustar_formato_celdas(output_file):
    wb = load_workbook(output_file)

    # Recorrer solo la hoja 'Detalle Marcación'
    worksheet = wb['Detalle Marcación']
    columnas = [cell.value for cell in worksheet[1]]

    marca_ing_col = columnas.index('Marca Ing')
    programada_entrada_col = columnas.index('Programada Entrada')

    for row in worksheet.iter_rows(min_row=2):
        marca_ing = row[marca_ing_col].value  # 'Marca Ing'
        programada_entrada = row[programada_entrada_col].value  # 'Programada Entrada'
        
        marca_ing_minutos = tiempo_a_minutos(marca_ing)
        programada_entrada_minutos = tiempo_a_minutos(programada_entrada)

        if marca_ing_minutos is not None and programada_entrada_minutos is not None:
            if marca_ing_minutos > programada_entrada_minutos:
                celda_marca_ing = row[marca_ing_col]
                celda_marca_ing.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo Amarillo
                celda_marca_ing.font = Font(color="FF0000", bold=True)  # Texto en color rojo y en negrita

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    wb.save(output_file)

def guardar_en_excel(data, output_file, pdf_path, excel_path, num_workers=4):
    start_time_total = time.time()
    print(f"Inicio de guardado en Excel: {output_file}")
    
    start_time_detalle = time.time()
    df = pd.DataFrame(data)
    df = agregar_columna_semana(df)
    df = agregar_columna_tipo_jornada(df, excel_path)
    df_sin_resumen, resumen_semanal = extraer_resumen_semanal(df)
    end_time_detalle = time.time()
    print(f"Creación de datos para 'Detalle Marcación' y 'Resumen Semanal' completada en {end_time_detalle - start_time_detalle:.2f} segundos.")

    start_time_saldo = time.time()
    saldo_horas_df = extraer_saldo_horas_parallel(pdf_path, num_workers=num_workers)
    saldo_horas_df = agregar_columna_tipo_jornada(saldo_horas_df, excel_path)

    end_time_saldo = time.time()
    if saldo_horas_df is not None:
        print(f"Creación de datos para 'Saldo de Horas' completada en {end_time_saldo - start_time_saldo:.2f} segundos.")
    else:
        print("No se encontró información para 'Saldo de Horas'.")

    start_time_guardado = time.time()
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        df_sin_resumen.to_excel(writer, sheet_name="Detalle Marcación", index=False)
    
        if not resumen_semanal.empty:
            resumen_semanal.to_excel(writer, sheet_name="Resumen Semanal", index=False)
    
        if saldo_horas_df is not None:
            saldo_horas_df.to_excel(writer, sheet_name="Saldo de Horas", index=False)
    
    end_time_guardado = time.time()
    print(f"Guardado en archivo Excel completado en {end_time_guardado - start_time_guardado:.2f} segundos.")
    print("Ajustando el formato del excel")
    ajustar_formato_celdas(output_file)

    end_time_total = time.time()
    print(f"Guardado en Excel total completado en {end_time_total - start_time_total:.2f} segundos.")
